"""Order Management — 订单管理 API（Order Router 模块）"""
import uuid
import random
from fastapi import APIRouter, Query, HTTPException, UploadFile, File
from sqlalchemy import func
from sqlalchemy.orm import joinedload

from app.core.database import SessionLocal
from app.models.order import Order, OrderItem, OrderRouting
from app.models.facility import Facility

router = APIRouter()

VALID_STATUSES = [
    "pending", "paid", "processing", "printed",
    "packed", "shipped", "completed", "cancelled",
]

def gen_uuid():
    return str(uuid.uuid4())


# ── GET /orders/stats — 统计各状态数量（必须在 /{order_id} 之前）──

@router.get("/stats")
def order_stats():
    db = SessionLocal()
    try:
        rows = (
            db.query(Order.status, func.count(Order.id))
            .group_by(Order.status)
            .all()
        )
        status_map = {s: c for s, c in rows}
        total = sum(status_map.values())

        return {
            "total": total,
            "pending": status_map.get("pending", 0),
            "paid": status_map.get("paid", 0),
            "processing": status_map.get("processing", 0),
            "printed": status_map.get("printed", 0),
            "packed": status_map.get("packed", 0),
            "shipped": status_map.get("shipped", 0),
            "completed": status_map.get("completed", 0),
            "cancelled": status_map.get("cancelled", 0),
            "by_status": status_map,
        }
    finally:
        db.close()


# ── GET /orders/ — 订单列表（分页 + 过滤） ───────────────────────

@router.get("/")
def list_orders(
    platform: str | None = Query(None),
    status: str | None = Query(None),
    channel: str | None = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
):
    db = SessionLocal()
    try:
        q = db.query(Order)

        if platform:
            q = q.filter(Order.platform == platform)
        if status:
            q = q.filter(Order.status == status)
        # channel 兼容旧参数，映射到 platform
        if channel and not platform:
            q = q.filter(Order.platform == channel)

        total = q.count()
        orders = (
            q.order_by(Order.created_at.desc())
            .offset((page - 1) * page_size)
            .limit(page_size)
            .all()
        )

        return {
            "data": [
                {
                    "id": o.id,
                    "platform": o.platform,
                    "platform_order_id": o.platform_order_id,
                    "customer_name": o.customer_name,
                    "status": o.status,
                    "total_amount": o.total_amount,
                    "currency": o.currency,
                    "is_rush": o.is_rush,
                    "created_at": str(o.created_at)[:19] if o.created_at else "",
                    "tracking_no": o.tracking_no,
                    "item_count": len(o.items) if o.items else 0,
                }
                for o in orders
            ],
            "total": total,
            "page": page,
            "page_size": page_size,
            "pages": (total + page_size - 1) // page_size,
        }
    finally:
        db.close()


# ── POST /orders/ — 手动创建订单 ─────────────────────────────────

@router.post("/")
def create_order(data: dict):
    db = SessionLocal()
    try:
        order = Order(
            id=gen_uuid(),
            tenant_id=data.get("tenant_id", "default"),
            platform=data.get("platform", "manual"),
            platform_order_id=data.get("platform_order_id") or data.get("order_no"),
            customer_name=data.get("customer_name"),
            customer_email=data.get("customer_email"),
            shipping_address=data.get("shipping_address"),
            phone=data.get("phone"),
            total_amount=float(data.get("total_amount", 0)),
            currency=data.get("currency", "USD"),
            status=data.get("status", "pending"),
            is_rush=data.get("is_rush", False),
            notes=data.get("notes"),
        )
        db.add(order)

        items_data = data.get("items", [])
        for item in items_data:
            db.add(OrderItem(
                id=gen_uuid(),
                order_id=order.id,
                tenant_id=order.tenant_id,
                sku=item.get("sku", ""),
                product_name=item.get("product_name"),
                qty=int(item.get("qty", 1)),
                unit_price=float(item.get("unit_price", 0)),
                production_status=item.get("production_status", "pending"),
            ))

        db.commit()
        db.refresh(order)
        return {
            "id": order.id,
            "platform": order.platform,
            "status": order.status,
            "total_amount": order.total_amount,
            "created_at": str(order.created_at)[:19] if order.created_at else "",
        }
    except Exception as e:
        db.rollback()
        raise HTTPException(400, str(e))
    finally:
        db.close()


# ── GET /orders/{order_id} — 订单详情（含 items + routing）───────

@router.get("/{order_id}")
def get_order(order_id: str):
    db = SessionLocal()
    try:
        order = (
            db.query(Order)
            .options(joinedload(Order.items), joinedload(Order.routing))
            .filter(Order.id == order_id)
            .first()
        )
        if not order:
            raise HTTPException(404, "Order not found")

        items = [
            {
                "id": i.id,
                "sku": i.sku,
                "product_name": i.product_name,
                "qty": i.qty,
                "unit_price": i.unit_price,
                "production_status": i.production_status,
            }
            for i in (order.items or [])
        ]

        routing = None
        if order.routing:
            r = order.routing
            routing = {
                "id": r.id,
                "proximity_score": r.proximity_score,
                "stock_score": r.stock_score,
                "capacity_score": r.capacity_score,
                "cost_score": r.cost_score,
                "total_score": r.total_score,
                "selected_facility_id": r.selected_facility_id,
                "status": r.status,
            }

        return {
            "id": order.id,
            "tenant_id": order.tenant_id,
            "platform": order.platform,
            "platform_order_id": order.platform_order_id,
            "customer_name": order.customer_name,
            "customer_email": order.customer_email,
            "shipping_address": order.shipping_address,
            "phone": order.phone,
            "total_amount": order.total_amount,
            "currency": order.currency,
            "status": order.status,
            "tracking_no": order.tracking_no,
            "carrier": order.carrier,
            "shipped_at": str(order.shipped_at)[:19] if order.shipped_at else "",
            "is_rush": order.is_rush,
            "notes": order.notes,
            "created_at": str(order.created_at)[:19] if order.created_at else "",
            "items": items,
            "routing": routing,
        }
    finally:
        db.close()


# ── PUT /orders/{order_id}/status — 更新状态 ─────────────────────

@router.put("/{order_id}/status")
def update_status(order_id: str, data: dict):
    db = SessionLocal()
    try:
        order = db.query(Order).filter(Order.id == order_id).first()
        if not order:
            raise HTTPException(404, "Order not found")

        new_status = data.get("status")
        if not new_status:
            raise HTTPException(400, "status is required")
        if new_status not in VALID_STATUSES:
            raise HTTPException(400, f"Invalid status. Must be one of: {VALID_STATUSES}")

        order.status = new_status
        db.commit()
        return {"id": order.id, "status": order.status}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(400, str(e))
    finally:
        db.close()


# ── POST /orders/{order_id}/route — 四维度评分路由 ───────────────

@router.post("/{order_id}/route")
def route_order(order_id: str, data: dict | None = None):
    """四维度评分：proximity / stock / capacity / cost，选最优印刷厂"""
    db = SessionLocal()
    try:
        order = db.query(Order).filter(Order.id == order_id).first()
        if not order:
            raise HTTPException(404, "Order not found")

        weights = (data or {}).get("weights", {})
        w_prox = weights.get("proximity", 0.25)
        w_stock = weights.get("stock", 0.25)
        w_cap = weights.get("capacity", 0.25)
        w_cost = weights.get("cost", 0.25)

        facilities = db.query(Facility).filter(Facility.is_active == True).all()
        if not facilities:
            raise HTTPException(400, "No active facilities available")

        results = []
        for f in facilities:
            # 1. proximity — 距离评分（基于经纬度归一化 0-100）
            if f.lat is not None and f.lng is not None:
                prox = max(0, min(100, (f.lat + 90) / 1.8 * 50 + (f.lng + 180) / 3.6 * 50))
            else:
                prox = round(random.uniform(40, 90), 1)

            # 2. stock — 库存充足度（产能利用率越低越充足）
            stock = round((1.0 - f.capacity_utilization) * 100, 1)

            # 3. capacity — 产能余量（利用率越低余量越大）
            cap = round((1.0 - f.capacity_utilization) * 100, 1)

            # 4. cost — 综合成本（cost_factor 越低评分越高）
            cost_factor = max(f.cost_factor, 0.01)
            cost = round(min(100, (1.0 / cost_factor) * 50), 1)

            total = round(
                w_prox * prox + w_stock * stock + w_cap * cap + w_cost * cost, 1
            )

            results.append({
                "facility_id": f.id,
                "facility_name": f.name,
                "proximity_score": prox,
                "stock_score": stock,
                "capacity_score": cap,
                "cost_score": cost,
                "total_score": total,
            })

        results.sort(key=lambda x: x["total_score"], reverse=True)
        best = results[0]

        # 写入/更新 OrderRouting
        existing = db.query(OrderRouting).filter(OrderRouting.order_id == order_id).first()
        if existing:
            existing.proximity_score = best["proximity_score"]
            existing.stock_score = best["stock_score"]
            existing.capacity_score = best["capacity_score"]
            existing.cost_score = best["cost_score"]
            existing.w_proximity = w_prox
            existing.w_stock = w_stock
            existing.w_capacity = w_cap
            existing.w_cost = w_cost
            existing.total_score = best["total_score"]
            existing.selected_facility_id = best["facility_id"]
            existing.status = "confirmed"
        else:
            routing = OrderRouting(
                id=gen_uuid(),
                order_id=order_id,
                tenant_id=order.tenant_id,
                proximity_score=best["proximity_score"],
                stock_score=best["stock_score"],
                capacity_score=best["capacity_score"],
                cost_score=best["cost_score"],
                w_proximity=w_prox,
                w_stock=w_stock,
                w_capacity=w_cap,
                w_cost=w_cost,
                total_score=best["total_score"],
                selected_facility_id=best["facility_id"],
                status="confirmed",
            )
            db.add(routing)

        if order.status == "pending":
            order.status = "processing"

        db.commit()

        return {
            "order_id": order_id,
            "selected_facility": best,
            "all_scores": results,
            "weights": {"proximity": w_prox, "stock": w_stock, "capacity": w_cap, "cost": w_cost},
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(400, str(e))
    finally:
        db.close()


# ── WMS Integration: 领星波次查询 ──
from pydantic import BaseModel

class WMSWaveRequest(BaseModel):
    waveNo: str

@router.post("/wms/wave-detail")
def wms_wave_detail(req: WMSWaveRequest):
    """从领星 WMS 拉取波次订单数据"""
    from app.services.wms_service import fetch_wave_detail
    try:
        result = fetch_wave_detail(req.waveNo)
        return result
    except Exception as e:
        return {"code": -1, "msg": str(e), "data": None}


class SyncRequest(BaseModel):
    waveNos: list

@router.post("/wms/sync-orders")
def sync_orders(req: SyncRequest):
    """从 WMS 批量同步波次订单到本地数据库"""
    from app.services.wms_service import sync_waves_to_db
    try:
        result = sync_waves_to_db(req.waveNos)
        return {"code":0, "data": result}
    except Exception as e:
        return {"code":-1, "msg": str(e)}

@router.get("/list/synced")
def list_synced_orders(ym: str = ""):
    """获取已同步的订单（支持年月筛选 YYYY-MM）"""
    from app.services.wms_service import get_synced_orders, get_available_months
    try:
        if not ym:
            months = get_available_months()
            if months: ym = months[0]  # latest month
        data = get_synced_orders(ym)
        data["available_months"] = get_available_months()
        data["current_ym"] = ym
        return {"code":0, "data": data}
    except Exception as e:
        return {"code":-1, "msg": str(e)}


@router.get("/delivery/kpi")
def delivery_kpi():
    """订单KPI统计：总单量/待处理/已打印/已拣货/已复核/已出库/已取消"""
    import sqlite3, os
    db_path = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(__file__)))), "data", "riichain.db")
    conn = sqlite3.connect(db_path)
    
    kpi = {}
    # Use order_no for order counts (wave_orders may have multiple SKUs per order)
    for row in conn.execute("SELECT custom_status, COUNT(DISTINCT order_no) FROM wave_orders GROUP BY custom_status"):
        kpi[row[0]] = row[1]
    
    # Total unique orders
    total = conn.execute("SELECT COUNT(DISTINCT order_no) FROM wave_orders").fetchone()[0]
    kpi["total"] = total
    
    # Available months  
    months = [r[0] for r in conn.execute("SELECT DISTINCT substr(synced_at, 1, 7) FROM wave_orders WHERE synced_at != '' ORDER BY synced_at DESC").fetchall()]
    
    conn.close()
    return {"code": 0, "data": {"kpi": kpi, "available_months": months}}

@router.get("/delivery/list")
def list_delivery_orders(ym: str = "", status: str = "", date_from: str = "", date_to: str = "", page: int = 1, page_size: int = 30):
    """获取出库单列表（从 wave_orders 查询，支持筛选 + 分页）"""
    import sqlite3, os, math
    db_path = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(__file__)))), "data", "riichain.db")
    conn = sqlite3.connect(db_path); conn.row_factory = sqlite3.Row
    
    where = []; params = []
    if ym:
        where.append("synced_at LIKE ?")
        params.append(ym + "%")
    if date_from:
        where.append("synced_at >= ?")
        params.append(date_from)
    if date_to:
        where.append("synced_at <= ?")
        params.append(date_to)
    if status:
        where.append("custom_status = ?")
        params.append(status)
    
    w = " WHERE " + " AND ".join(where) if where else ""
    
    # Get distinct orders for count
    count_sql = f"SELECT COUNT(DISTINCT order_no) FROM wave_orders{w}"
    total = int(conn.execute(count_sql, params).fetchone()[0])
    pages = max(1, math.ceil(total / page_size))
    
    # Get page of distinct orders
    offset = (page - 1) * page_size
    sql = f"""SELECT order_no, wave_no, tracking_no, carrier, status, 
               custom_status, synced_at,
               COUNT(*) as sku_count, SUM(qty) as total_qty,
               GROUP_CONCAT(DISTINCT sku_code) as skus
        FROM wave_orders{w} 
        GROUP BY order_no 
        ORDER BY synced_at DESC, order_no 
        LIMIT ? OFFSET ?"""
    rows = conn.execute(sql, params + [page_size, offset]).fetchall()
    orders = [dict(r) for r in rows]
    conn.close()
    
    return {"code": 0, "data": {"orders": orders, "total": total, "page": page, "pages": pages, "page_size": page_size}}

@router.post("/delivery/update-status")
def update_delivery_status(data: dict):
    """更新订单自定义状态（从 wave_orders 更新）"""
    import sqlite3, os
    db_path = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(__file__)))), "data", "riichain.db")
    conn = sqlite3.connect(db_path)
    order_no = data.get("order_no", "")
    new_status = data.get("custom_status", "")
    if not order_no or not new_status:
        return {"code": -1, "msg": "missing params"}
    
    # 更新 wave_orders 表
    conn.execute("UPDATE wave_orders SET custom_status = ? WHERE order_no = ?", (new_status, order_no))
    updated = conn.total_changes
    conn.commit()
    conn.close()
    return {"code": 0, "data": {"updated": updated, "order_no": order_no, "status": new_status}}
@router.post("/delivery/upload-excel")
async def upload_delivery_excel(file: UploadFile = File(...)):
    """上传出库单Excel，数据写入 wave_orders 表"""
    import openpyxl
    from io import BytesIO
    import time, uuid
    
    db_path = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(__file__)))), "data", "riichain.db")
    
    wb = openpyxl.load_workbook(BytesIO(await file.read()), read_only=True)
    ws = wb.active
    headers = [str(c.value or '') for c in next(ws.iter_rows(min_row=1, max_row=1))]
    
    cols = {}
    for i, h in enumerate(headers):
        h2 = h.lower()
        if 'outbound order' in h2: cols['outbound_no'] = i
        elif 'wave no' in h2: cols['wave_no'] = i
        elif 'reference order' in h2: cols['refer_order_no'] = i
        elif 'platform number' in h2: cols['platform_no'] = i
        elif 'shipping carrier' in h2: cols['carrier'] = i
        elif h2 == 'status': cols['status'] = i
        elif 'creation time' in h2: cols['creation_time'] = i
        elif 'outboundtime' in h2 or 'outbound time' in h2: cols['outbound_time'] = i
        elif 'total qty' in h2: cols['total_qty'] = i
        elif 'tracking no' in h2: cols['tracking_no'] = i
        elif h2 == 'sku': cols['sku'] = i
        elif 'product name' in h2: cols['product_name'] = i
        elif 'product type' in h2: cols['product_type'] = i
        elif 'outbound qty' in h2: cols['outbound_qty'] = i
    
    if 'outbound_no' not in cols: 
        return {"code": -1, "msg": "未找到Outbound Order No列"}
    
    conn = sqlite3.connect(db_path)
    today = time.strftime('%Y-%m-%d %H:%M:%S')
    count = 0
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        def g(k): 
            v = row[cols[k]] if k in cols and cols[k] < len(row) else ''
            return str(v).strip() if v else ''
        
        ono = g('outbound_no')
        if not ono: 
            continue
        
        ds = ono[6:12] if len(ono) >= 12 else ''
        od = '20' + ds[:2] + '-' + ds[2:4] + '-' + ds[4:6] if ds else ''
        
        try:
            # 检查是否已存在
            existing = conn.execute('SELECT id FROM wave_orders WHERE order_no = ?', (ono,)).fetchone()
            
            if existing:
                # 更新已有记录
                conn.execute("""UPDATE wave_orders SET 
                    wave_no = COALESCE(?, wave_no),
                    tracking_no = COALESCE(?, tracking_no),
                    carrier = COALESCE(?, carrier),
                    product_type = COALESCE(?, product_type),
                    qty = COALESCE(?, qty),
                    sku_code = COALESCE(?, sku_code),
                    synced_at = ?
                    WHERE order_no = ?""",
                    (g('wave_no'), g('tracking_no'), g('carrier'),
                     g('product_type'), 
                     int(float(g('outbound_qty'))) if g('outbound_qty') else None,
                     g('sku'), today, ono))
            else:
                # 插入新记录
                new_id = str(uuid.uuid4())
                conn.execute("""INSERT INTO wave_orders 
                    (id, order_no, outbound_no, wave_no, sku_code, qty,
                     product_type, tracking_no, carrier, synced_at, status)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                    (new_id, ono, ono, g('wave_no'), g('sku'),
                     int(float(g('outbound_qty')) if g('outbound_qty') else 1),
                     g('product_type') or 'Customization',
                     g('tracking_no'), g('carrier'), today, '待处理'))
            count += 1
        except Exception as e:
            print(f"插入失败 {ono}: {e}")
            continue
    
    # 更新 custom_status
    conn.execute("""UPDATE wave_orders SET custom_status = 'pending' 
        WHERE (wave_no = '' OR wave_no IS NULL) 
        AND (status != '8' OR status IS NULL) 
        AND (custom_status = '' OR custom_status IS NULL)""")
    
    conn.execute("""UPDATE wave_orders SET custom_status = 'printed' 
        WHERE wave_no != '' AND wave_no IS NOT NULL 
        AND (custom_status = '' OR custom_status IS NULL)""")
    
    conn.execute("UPDATE wave_orders SET custom_status = 'cancelled' WHERE status = '8'")
    
    conn.commit()
    conn.close()
    wb.close()
    
    return {"code": 0, "data": {"imported": count, "msg": f"成功导入{count}条记录"}}


# ── POST /api/v1/orders/wms/sync-delivery ──

class DeliverySyncRequest(BaseModel):
    """同步出库单请求"""
    date: str = ""  # YYYY-MM-DD，默认为今天
    overwrite: bool = False  # 是否覆盖已存在的订单
    fetch_details: bool = False  # 是否自动获取订单详情（SKU列表、物流信息）
    wh_code: str = ""  # 仓库代码，默认为 TXMISSOURI


@router.post("/wms/sync-delivery")
def sync_delivery_orders(req: DeliverySyncRequest):
    """
    从领星WMS同步出库单到RiiChain
    使用 /openapi/v2/delivery/page API 获取订单列表
    如果 fetch_details=True，还会调用 /openapi/v2/delivery/detail 获取订单详情（SKU、物流信息）
    """
    from app.services.wms_service import sync_delivery_orders_by_date
    
    try:
        result = sync_delivery_orders_by_date(
            date_str=req.date,
            overwrite=req.overwrite,
            wh_code=req.wh_code,
            fetch_details=req.fetch_details
        )
        
        msg = f"同步完成：{result['synced_orders']}个订单"
        if result.get('skipped_orders', 0) > 0:
            msg += f"，跳过{result['skipped_orders']}个"
        if result.get('detail_result'):
            detail = result['detail_result']
            msg += f"，获取详情：{detail['synced_orders']}个订单，{detail['synced_skus']}个SKU"
        
        return {
            "code": 0,
            "msg": msg,
            "data": result
        }
    except Exception as e:
        return {
            "code": -1,
            "msg": f"同步失败：{str(e)}",
            "data": None
        }


@router.get("/wms/sync-delivery/preview")
def preview_delivery_sync(date: str = ""):
    """
    预览将要同步的出库单（不写入数据库）
    用于确认API是否正常、数据格式是否正确
    """
    from app.services.wms_service import fetch_delivery_page
    
    if not date:
        import time
        date = time.strftime("%Y-%m-%d")
    
    try:
        result = fetch_delivery_page(date, page_no=1, page_size=10)
        if result.get("code") != 0:
            return result
        
        return {
            "code": 0,
            "msg": "预览成功",
            "data": {
                "date": date,
                "total": result["data"].get("total", 0),
                "page_size": result["data"].get("pageSize", 0),
                "current_page": result["data"].get("page", 1),
                "sample_orders": result["data"].get("list", [])[:5]  # 返回前5条作为样本
            }
        }
    except Exception as e:
        return {
            "code": -1,
            "msg": f"预览失败：{str(e)}",
            "data": None
        }


# ── POST /wms/sync-delivery-details ──

class DeliveryDetailSyncRequest(BaseModel):
    """同步出库单详情请求"""
    outbound_nos: list = []  # 指定要同步的订单号列表，为空则同步所有没有SKU信息的订单
    date: str = ""  # 如果指定日期，则先同步该日期的订单列表，再获取详情


@router.post("/wms/sync-delivery-details")
def sync_delivery_details(req: DeliveryDetailSyncRequest):
    """
    同步出库单详情（SKU列表、物流信息）
    使用 /openapi/v2/delivery/detail API
    
    使用场景：
    1. 订单已同步（基本信息），但需要获取SKU详情和物流信息
    2. 指定 outbound_nos 同步特定订单的详情
    3. 不指定 outbound_nos 则同步所有没有SKU信息的订单
    """
    from app.services.wms_service import sync_order_details, sync_delivery_orders_by_date
    import sqlite3, os
    
    try:
        outbound_nos = req.outbound_nos
        
        # 如果指定了日期，先同步订单列表
        if req.date:
            sync_result = sync_delivery_orders_by_date(
                date_str=req.date,
                fetch_details=False  # 先不同步详情，后面统一处理
            )
            # 获取当天同步的订单号
            if not outbound_nos:
                db_path = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(__file__)))), "data", "riichain.db")
                conn = sqlite3.connect(db_path)
                rows = conn.execute(
                    "SELECT DISTINCT order_no FROM wave_orders WHERE synced_at LIKE ? AND (sku_code IS NULL OR sku_code = '')",
                    (req.date + "%",)
                ).fetchall()
                conn.close()
                outbound_nos = [r[0] for r in rows]
        
        # 如果没有指定订单号，则查询所有没有SKU信息的订单
        if not outbound_nos:
            db_path = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(__file__)))), "data", "riichain.db")
            conn = sqlite3.connect(db_path)
            rows = conn.execute(
                "SELECT DISTINCT order_no FROM wave_orders WHERE sku_code IS NULL OR sku_code = ''"
            ).fetchall()
            conn.close()
            outbound_nos = [r[0] for r in rows]
        
        if not outbound_nos:
            return {
                "code": 0,
                "msg": "没有需要同步详情的订单",
                "data": {"synced_orders": 0, "synced_skus": 0, "errors": []}
            }
        
        # 同步订单详情
        result = sync_order_details(outbound_nos)
        
        return {
            "code": 0,
            "msg": f"详情同步完成：{result['synced_orders']}个订单，{result['synced_skus']}个SKU",
            "data": result
        }
    except Exception as e:
        return {
            "code": -1,
            "msg": f"详情同步失败：{str(e)}",
            "data": None
        }
