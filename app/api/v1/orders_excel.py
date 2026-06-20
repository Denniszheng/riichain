"""Excel upload endpoint for delivery orders"""
import time
from fastapi import APIRouter, UploadFile, File
import openpyxl, sqlite3, os
from io import BytesIO

router = APIRouter()

@router.post("/delivery/upload-excel")
async def upload_delivery_excel(file: UploadFile = File(...)):
    db_path=os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(__file__)))),"data","riichain.db")

    wb=openpyxl.load_workbook(BytesIO(await file.read()),read_only=True)
    ws=wb.active
    headers=[str(c.value or '') for c in next(ws.iter_rows(min_row=1,max_row=1))]

    cols={}
    for i,h in enumerate(headers):
        h2=h.lower()
        if 'outbound order' in h2: cols['outbound_no']=i
        elif 'wave no' in h2: cols['wave_no']=i
        elif 'reference order' in h2: cols['refer_order_no']=i
        elif 'platform number' in h2: cols['platform_no']=i
        elif 'shipping carrier' in h2: cols['carrier']=i
        elif h2=='status': cols['status']=i
        elif 'creation time' in h2: cols['creation_time']=i
        elif 'outboundtime' in h2 or 'outbound time' in h2: cols['outbound_time']=i
        elif 'total qty' in h2: cols['total_qty']=i
        elif 'tracking no' in h2: cols['tracking_no']=i
        elif h2=='sku': cols['sku']=i
        elif 'product name' in h2: cols['product_name']=i
        elif 'product type' in h2: cols['product_type']=i
        elif 'outbound qty' in h2: cols['outbound_qty']=i

    if 'outbound_no' not in cols: return {"code":-1,"msg":"未找到Outbound Order No列"}

    conn=sqlite3.connect(db_path)
    conn.execute("CREATE TABLE IF NOT EXISTS delivery_orders (id INTEGER PRIMARY KEY AUTOINCREMENT, outbound_no TEXT, wave_no TEXT, refer_order_no TEXT, platform_no TEXT, carrier TEXT, status TEXT, creation_time TEXT, outbound_time TEXT, total_qty INTEGER, tracking_no TEXT, sku TEXT, product_name TEXT, product_type TEXT, outbound_qty INTEGER, order_date TEXT, synced_at TEXT, custom_status TEXT DEFAULT '')")
    try: conn.execute("ALTER TABLE delivery_orders ADD COLUMN custom_status TEXT DEFAULT ''")
    except: pass

    today=time.strftime('%Y-%m-%d %H:%M:%S'); count=0
    for row in ws.iter_rows(min_row=2,values_only=True):
        def g(k): v=row[cols[k]] if k in cols and cols[k]<len(row) else ''; return str(v).strip() if v else ''
        ono=g('outbound_no')
        if not ono: continue
        ds=ono[6:12] if len(ono)>=12 else ''
        od='20'+ds[:2]+'-'+ds[2:4]+'-'+ds[4:6] if ds else ''
        try:
            conn.execute("INSERT OR REPLACE INTO delivery_orders (outbound_no,wave_no,refer_order_no,platform_no,carrier,status,creation_time,outbound_time,total_qty,tracking_no,sku,product_name,product_type,outbound_qty,order_date,synced_at) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                (ono,g('wave_no'),g('refer_order_no'),g('platform_no'),g('carrier'),g('status'),
                 g('creation_time'),g('outbound_time'),int(float(g('total_qty')) if g('total_qty') else 0),
                 g('tracking_no'),g('sku'),g('product_name'),g('product_type'),
                 int(float(g('outbound_qty')) if g('outbound_qty') else 0),od,today)); count+=1
        except: pass

    conn.execute("UPDATE delivery_orders SET custom_status='pending' WHERE (wave_no='' OR wave_no IS NULL) AND status!='8' AND custom_status=''")
    conn.execute("UPDATE delivery_orders SET custom_status='printed' WHERE wave_no!='' AND wave_no IS NOT NULL AND custom_status=''")
    conn.execute("UPDATE delivery_orders SET custom_status='cancelled' WHERE status='8'")
    conn.commit(); conn.close()
    wb.close()

    return {"code":0,"data":{"imported":count,"msg":"成功导入"+str(count)+"条记录"}}
