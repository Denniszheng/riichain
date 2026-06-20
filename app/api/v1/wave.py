"""Wave & Fulfillment — 波次履约 API（迁移自 pod-platform）"""
import io
from fastapi import APIRouter, UploadFile, File, Form, HTTPException
from sqlalchemy.orm import Session
from openpyxl import load_workbook
from app.core.database import SessionLocal
from app.models.wave import WaveImport, WaveOrder, MaterialInventory, StockTransaction

router = APIRouter()

FIELD_PATTERNS = {
    "order_no":    ["Outbound Order No", "出库单号", "订单号", "Order No"],
    "wave_no":     ["Wave No.", "波次号", "关联波次号", "Wave"],
    "tracking_no": ["Tracking No.", "快递单号", "运单号", "Tracking"],
    "sku_code":    ["SKU", "sku", "Sku"],
    "qty":         ["Outbound Qty", "出库数量", "数量", "Qty", "Quantity"],
    "product_type":["Product Type", "产品分类", "产品类型", "Type", "Category"],
    "status":      ["状态", "Status"],
    "order_date":  ["来单日期", "订单日期", "Order Date"],
    "ship_date":   ["出库日期", "发货日期", "Ship Date"],
}


# ── Import Management ────────────────────────────────────────────

@router.get("/imports")
def list_wave_imports(skip: int = 0, limit: int = 15, status: str | None = None):
    """列出所有 Excel 导入批次（支持 status 过滤）"""
    db: Session = SessionLocal()
    try:
        q = db.query(WaveImport).order_by(WaveImport.created_at.desc())
        if status:
            q = q.filter(WaveImport.status == status)
        total = q.count()
        imports = q.offset(skip).limit(limit).all()
        return {
            "data": [
                {
                    "id": imp.id,
                    "filename": imp.filename,
                    "total_rows": imp.total_rows,
                    "total_orders": imp.total_orders or 0,
                    "total_skus": imp.total_skus or 0,
                    "total_waves": imp.total_waves or 0,
                    "status": imp.status or "待确认",
                    "created_at": str(imp.created_at)[:19] if imp.created_at else "",
                }
                for imp in imports
            ],
            "total": total,
        }
    finally:
        db.close()


@router.post("/imports/upload")
def upload_wave_file(file: UploadFile = File(...), tenant_id: str = Form("default")):
    """上传 Excel → 解析 wave → 写入 DB"""
    wb = load_workbook(io.BytesIO(file.file.read()), read_only=True)
    ws = wb.active
    header = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]

    col_map = {}
    for c, hdr in enumerate(header):
        for field, patterns in FIELD_PATTERNS.items():
            if any(p in hdr for p in patterns):
                col_map[field] = c
                break

    rows = []
    for row_cells in ws.iter_rows(min_row=2):
        vals = [str(c.value or "").strip() for c in row_cells]
        order_no = vals[col_map["order_no"]] if "order_no" in col_map else ""
        wave_no = vals[col_map["wave_no"]] if "wave_no" in col_map else ""
        if not order_no or not wave_no:
            continue
        sku = vals[col_map["sku_code"]] if "sku_code" in col_map else (vals[3] if len(vals) > 3 else "")
        try:
            qty = int(float(vals[col_map["qty"]] or 0)) if "qty" in col_map else (int(float(vals[4] or 0)) if len(vals) > 4 else 1)
        except (ValueError, IndexError):
            qty = 1
        rows.append({
            "order_no": order_no,
            "wave_no": wave_no,
            "tracking_no": vals[col_map["tracking_no"]] if "tracking_no" in col_map else (vals[2] if len(vals) > 2 else ""),
            "sku_code": sku,
            "qty": qty,
            "product_type": vals[col_map["product_type"]] if "product_type" in col_map else (vals[5] if len(vals) > 5 else "Standard"),
        })

    if not rows:
        raise HTTPException(400, "Excel 文件中无有效数据行")

    waves = list(set(r["wave_no"] for r in rows))
    db: Session = SessionLocal()
    try:
        imp = WaveImport(
            tenant_id=tenant_id,
            filename=file.filename,
            total_rows=len(rows),
            total_orders=len(set(r["order_no"] for r in rows)),
            total_skus=len(set(r["sku_code"] for r in rows)),
            total_waves=len(waves),
        )
        db.add(imp)
        db.flush()
        for r in rows:
            db.add(WaveOrder(import_id=imp.id, tenant_id=tenant_id, **r))
        db.commit()
        return {"import_id": imp.id, "total_rows": len(rows), "total_waves": len(waves), "waves": sorted(waves)}
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


@router.post("/imports/{import_id}/confirm")
def confirm_wave_import(import_id: str):
    """确认导入，状态 → 已确认"""
    db: Session = SessionLocal()
    try:
        imp = db.query(WaveImport).filter(WaveImport.id == import_id).first()
        if not imp:
            raise HTTPException(404, "导入批次不存在")
        imp.status = "已确认"
        db.commit()
        return {"import_id": import_id, "status": "已确认"}
    except HTTPException:
        raise
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


# ── Wave Orders ──────────────────────────────────────────────────

@router.get("/orders")
def list_wave_orders(
    wave_no: str | None = None,
    status: str | None = None,
    skip: int = 0,
    limit: int = 100,
):
    """列出 WaveOrder（支持 wave_no/status 过滤，分页）"""
    db: Session = SessionLocal()
    try:
        q = db.query(WaveOrder)
        if wave_no:
            q = q.filter(WaveOrder.wave_no == wave_no)
        if status:
            q = q.filter(WaveOrder.status == status)
        total = q.count()
        orders = q.offset(skip).limit(limit).all()
        return {
            "data": [
                {
                    "id": o.id,
                    "order_no": o.order_no,
                    "wave_no": o.wave_no,
                    "tracking_no": o.tracking_no,
                    "sku_code": o.sku_code,
                    "qty": o.qty,
                    "product_type": o.product_type,
                    "status": o.status,
                    "order_date": o.order_date,
                    "ship_date": o.ship_date,
                }
                for o in orders
            ],
            "total": total,
        }
    finally:
        db.close()


@router.put("/orders/{order_id}/status")
def update_wave_order_status(order_id: str, data: dict):
    """更新 WaveOrder 状态"""
    db: Session = SessionLocal()
    try:
        order = db.query(WaveOrder).filter(WaveOrder.id == order_id).first()
        if not order:
            raise HTTPException(404, "订单不存在")
        new_status = data.get("status")
        valid = ["待处理", "待创波次", "待打印", "待分货", "已分货", "已发货"]
        if new_status not in valid:
            raise HTTPException(400, f"无效状态，允许：{valid}")
        order.status = new_status
        db.commit()
        return {"id": order_id, "status": new_status}
    except HTTPException:
        raise
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


# ── Wave Detail ──────────────────────────────────────────────────

@router.get("/materials/list")
def list_materials():
    """列出所有耗材库存（亚克力板等）"""
    db: Session = SessionLocal()
    try:
        materials = db.query(MaterialInventory).order_by(MaterialInventory.board_size).all()
        return {
            "data": [
                {
                    "id": m.id,
                    "board_size": m.board_size,
                    "qty_on_hand": m.qty_on_hand,
                    "min_stock": m.min_stock,
                    "supplier": m.supplier,
                    "is_low": m.qty_on_hand < m.min_stock,
                }
                for m in materials
            ]
        }
    finally:
        db.close()


@router.post("/materials/deduct")
def deduct_material(data: dict):
    """排版后扣减耗材库存"""
    db: Session = SessionLocal()
    try:
        board_size = data.get("board_size", "480x480")
        qty = data.get("qty", 1)
        mat = db.query(MaterialInventory).filter(MaterialInventory.board_size == board_size).first()
        if not mat:
            raise HTTPException(404, f"材料 {board_size} 不存在")
        if mat.qty_on_hand < qty:
            raise HTTPException(400, f"{board_size} 库存不足：需要 {qty}，当前 {mat.qty_on_hand}")
        mat.qty_on_hand -= qty
        db.add(StockTransaction(
            tenant_id="default", board_size=board_size,
            transaction_type="layout_consume", quantity_change=-qty,
            quantity_after=mat.qty_on_hand,
            reference=data.get("layout_ref", ""),
        ))
        db.commit()
        return {"board_size": board_size, "remaining": mat.qty_on_hand}
    except HTTPException:
        raise
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


@router.get("/{wave_no}")
def get_wave_detail(wave_no: str):
    """查看一个 Wave 下的所有订单 + SKU"""
    db: Session = SessionLocal()
    try:
        orders = db.query(WaveOrder).filter(WaveOrder.wave_no == wave_no).order_by(WaveOrder.order_no).all()
        if not orders:
            raise HTTPException(404, "波次不存在")

        from collections import defaultdict
        order_map = defaultdict(list)
        for o in orders:
            order_map[o.order_no].append({"sku_code": o.sku_code, "qty": o.qty, "product_type": o.product_type})

        return {
            "wave_no": wave_no,
            "order_count": len(order_map),
            "total_qty": sum(o.qty for o in orders),
            "orders": [{"order_no": k, "skus": v} for k, v in order_map.items()],
        }
    except HTTPException:
        raise
    finally:
        db.close()
